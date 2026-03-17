from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
from tqdm import tqdm
import os


def aplicar_cores_serie(chart, cores):
    serie = chart.series[0]
    pontos = []

    for idx, cor in enumerate(cores):
        ponto = DataPoint(idx=idx)
        ponto.graphicalProperties.solidFill = cor
        ponto.graphicalProperties.line.solidFill = cor
        pontos.append(ponto)

    serie.dPt = pontos


def criar_dashboard(workbook, totais_mes, totais_ano, mes_nome, ano_principal):
    dashboard = workbook.create_sheet("Dashboard")
    resumo = workbook.create_sheet("Resumo")
    resumo.sheet_state = "hidden"

    dashboard.sheet_view.showGridLines = False

    # Cores
    fundo = "0F172A"
    fundo_bloco = "1E293B"
    azul = "3B82F6"
    verde = "22C55E"
    branco = "FFFFFF"
    cinza = "CBD5E1"
    borda_cor = "334155"

    paleta = [
        "4F46E5",
        "F59E0B",
        "A855F7",
        "22C55E",
        "EF4444",
        "06B6D4",
        "EAB308",
        "EC4899",
        "14B8A6",
        "8B5CF6",
    ]

    categorias = sorted(set(list(totais_mes.keys()) + list(totais_ano.keys())))

    if not categorias:
        categorias = ["Sem categoria"]
        totais_mes = {"Sem categoria": 0}
        totais_ano = {"Sem categoria": 0}

    total_mes = sum(totais_mes.values())
    total_ano = sum(totais_ano.values())

    # Fundo geral
    for row in range(1, 40):
        for col in range(1, 15):
            dashboard.cell(row=row, column=col).fill = PatternFill(
                fill_type="solid", fgColor=fundo
            )

    # Larguras
    larguras = {
        "A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 4,
        "H": 16, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16
    }
    for coluna, largura in larguras.items():
        dashboard.column_dimensions[coluna].width = largura

    # Altura das linhas
    for row in range(1, 40):
        dashboard.row_dimensions[row].height = 24

    borda = Border(
        left=Side(style="thin", color=borda_cor),
        right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor),
        bottom=Side(style="thin", color=borda_cor),
    )

    # Blocos
    blocos = [
        ("A1:F16"),
        ("H1:M16"),
        ("A18:M32"),
    ]

    for intervalo in blocos:
        inicio, fim = intervalo.split(":")
        col_i, row_i = inicio[0], int(inicio[1:])
        col_f, row_f = fim[0], int(fim[1:])

        for row in range(row_i, row_f + 1):
            for col in range(ord(col_i), ord(col_f) + 1):
                celula = dashboard.cell(row=row, column=col - 64)
                celula.fill = PatternFill(fill_type="solid", fgColor=fundo_bloco)
                celula.border = borda

    # Títulos
    dashboard.merge_cells("A1:F2")
    dashboard["A1"] = "Despesas por Categoria no Mês"
    dashboard["A1"].font = Font(bold=True, color=branco, size=16)
    dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center")

    dashboard.merge_cells("H1:M2")
    dashboard["H1"] = "Despesas por Categoria no Ano"
    dashboard["H1"].font = Font(bold=True, color=branco, size=16)
    dashboard["H1"].alignment = Alignment(horizontal="center", vertical="center")

    dashboard.merge_cells("A18:M19")
    dashboard["A18"] = "Resumo por Categoria"
    dashboard["A18"].font = Font(bold=True, color=branco, size=16)
    dashboard["A18"].alignment = Alignment(horizontal="center", vertical="center")

    dashboard.merge_cells("A15:F16")
    dashboard["A15"] = f"Total do mês: R$ {total_mes:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    dashboard["A15"].font = Font(bold=True, color=cinza, size=12)
    dashboard["A15"].alignment = Alignment(horizontal="center", vertical="center")

    dashboard.merge_cells("H15:M16")
    dashboard["H15"] = f"Total do ano: R$ {total_ano:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    dashboard["H15"].font = Font(bold=True, color=cinza, size=12)
    dashboard["H15"].alignment = Alignment(horizontal="center", vertical="center")

    # ABA RESUMO
    resumo["A1"] = "Categoria"
    resumo["B1"] = "Total Mês"
    resumo["C1"] = "% Mês"
    resumo["D1"] = "Total Ano"
    resumo["E1"] = "% Ano"

    for celula in resumo[1]:
        celula.font = Font(bold=True)

    for idx, categoria in enumerate(categorias, start=2):
        valor_mes = totais_mes.get(categoria, 0)
        valor_ano = totais_ano.get(categoria, 0)
        perc_mes = (valor_mes / total_mes) if total_mes > 0 else 0
        perc_ano = (valor_ano / total_ano) if total_ano > 0 else 0

        resumo[f"A{idx}"] = categoria
        resumo[f"B{idx}"] = valor_mes
        resumo[f"C{idx}"] = perc_mes
        resumo[f"D{idx}"] = valor_ano
        resumo[f"E{idx}"] = perc_ano

    # GRÁFICO ROSCA MÊS
    chart_mes = DoughnutChart()
    chart_mes.holeSize = 68
    chart_mes.varyColors = False
    chart_mes.legend.position = "r"
    chart_mes.height = 7
    chart_mes.width = 7

    dados_mes = Reference(resumo, min_col=2, min_row=2, max_row=1 + len(categorias))
    cats_mes = Reference(resumo, min_col=1, min_row=2, max_row=1 + len(categorias))
    chart_mes.add_data(dados_mes, titles_from_data=False)
    chart_mes.set_categories(cats_mes)
    aplicar_cores_serie(chart_mes, paleta[:len(categorias)])
    dashboard.add_chart(chart_mes, "A3")

    # GRÁFICO BARRAS ANO
    chart_ano = BarChart()
    chart_ano.type = "bar"
    chart_ano.style = 10
    chart_ano.height = 8
    chart_ano.width = 9
    chart_ano.y_axis.title = "Categorias"
    chart_ano.x_axis.title = "Total"
    chart_ano.legend = None

    dados_ano = Reference(resumo, min_col=4, min_row=1, max_row=1 + len(categorias))
    cats_ano = Reference(resumo, min_col=1, min_row=2, max_row=1 + len(categorias))
    chart_ano.add_data(dados_ano, titles_from_data=True)
    chart_ano.set_categories(cats_ano)
    dashboard.add_chart(chart_ano, "H3")

    # Informações de base fora da célula mesclada
    dashboard.merge_cells("H20:J20")
    dashboard["H20"] = f"Mês base: {mes_nome.capitalize()}"
    dashboard["H20"].font = Font(bold=True, color=verde)
    dashboard["H20"].alignment = Alignment(horizontal="center", vertical="center")
    dashboard["H20"].fill = PatternFill(fill_type="solid", fgColor=fundo_bloco)
    dashboard["H20"].border = borda

    dashboard.merge_cells("K20:M20")
    dashboard["K20"] = f"Ano base: {ano_principal}"
    dashboard["K20"].font = Font(bold=True, color=verde)
    dashboard["K20"].alignment = Alignment(horizontal="center", vertical="center")
    dashboard["K20"].fill = PatternFill(fill_type="solid", fgColor=fundo_bloco)
    dashboard["K20"].border = borda

    # TABELA RESUMO
    inicio_tabela = 22
    cabecalhos = ["Categoria", "Total Mês", "% Mês", "Total Ano", "% Ano"]

    for col, cabecalho in enumerate(cabecalhos, start=1):
        celula = dashboard.cell(row=inicio_tabela, column=col, value=cabecalho)
        celula.font = Font(bold=True, color=branco)
        celula.fill = PatternFill(fill_type="solid", fgColor=azul)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    for idx, categoria in enumerate(categorias, start=inicio_tabela + 1):
        valor_mes = totais_mes.get(categoria, 0)
        valor_ano = totais_ano.get(categoria, 0)
        perc_mes = (valor_mes / total_mes) if total_mes > 0 else 0
        perc_ano = (valor_ano / total_ano) if total_ano > 0 else 0

        linha = [
            categoria,
            valor_mes,
            perc_mes,
            valor_ano,
            perc_ano,
        ]

        for col, valor in enumerate(linha, start=1):
            celula = dashboard.cell(row=idx, column=col, value=valor)
            celula.alignment = Alignment(horizontal="center", vertical="center")
            celula.border = borda
            celula.fill = PatternFill(fill_type="solid", fgColor=fundo_bloco)
            celula.font = Font(color=branco)

            if col in (2, 4):
                celula.number_format = 'R$ #,##0.00'
            elif col in (3, 5):
                celula.number_format = '0.0%'


def exportar_gastos_excel(gastos):
    workbook = Workbook()
    pagina = workbook.active
    pagina.title = "Gastos"

    cabecalhos = ["Nome", "Valor", "Categoria", "Data", "Descrição"]
    pagina.append(cabecalhos)

    meses_pt = {
        "January": "janeiro",
        "February": "fevereiro",
        "March": "março",
        "April": "abril",
        "May": "maio",
        "June": "junho",
        "July": "julho",
        "August": "agosto",
        "September": "setembro",
        "October": "outubro",
        "November": "novembro",
        "December": "dezembro",
    }

    cor_cabecalho = PatternFill(fill_type="solid", fgColor="1F4E78")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")
    cor_linha_par = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cor_linha_impar = PatternFill(fill_type="solid", fgColor="FFFFFF")
    cor_total = PatternFill(fill_type="solid", fgColor="C6E0B4")
    fonte_total = Font(bold=True, color="000000")

    borda_fina = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for celula in pagina[1]:
        celula.font = fonte_cabecalho
        celula.fill = cor_cabecalho
        celula.alignment = alinhamento_centralizado
        celula.border = borda_fina

    meses = []
    total_gastos = 0
    registros_normalizados = []

    for gasto in tqdm(
        gastos,
        desc="Exportando gastos",
        unit="gasto",
        ncols=160
    ):
        if isinstance(gasto, dict):
            data = gasto.get("data", "")
            nome = gasto.get("nome", "")
            descricao = gasto.get("descricao", "")
            categoria = gasto.get("categoria", "")
            valor = gasto.get("valor", 0)
        else:
            data = gasto.data
            nome = gasto.nome
            descricao = gasto.descricao
            categoria = gasto.categoria
            valor = gasto.valor

        data_obj = None

        try:
            data_obj = datetime.strptime(data, "%d/%m/%Y")
            mes_en = data_obj.strftime("%B")
            meses.append(meses_pt.get(mes_en, mes_en))
        except ValueError:
            pass

        pagina.append([nome, float(valor), categoria, data, descricao])
        total_gastos += float(valor)

        registros_normalizados.append(
            {
                "nome": nome,
                "valor": float(valor),
                "categoria": categoria if categoria else "Sem categoria",
                "data": data,
                "data_obj": data_obj,
                "descricao": descricao,
            }
        )

    for linha in pagina.iter_rows(min_row=2, max_row=pagina.max_row, min_col=1, max_col=5):
        numero_linha = linha[0].row
        preenchimento = cor_linha_par if numero_linha % 2 == 0 else cor_linha_impar

        for celula in linha:
            celula.fill = preenchimento
            celula.border = borda_fina
            celula.alignment = alinhamento_centralizado

            if celula.column == 2:
                celula.number_format = 'R$ #,##0.00'

    linha_total = pagina.max_row + 1
    pagina.cell(row=linha_total, column=1, value="TOTAL")
    pagina.cell(row=linha_total, column=2, value=total_gastos)

    for col in range(1, 6):
        celula = pagina.cell(row=linha_total, column=col)
        celula.fill = cor_total
        celula.font = fonte_total
        celula.border = borda_fina
        celula.alignment = alinhamento_centralizado

    pagina.cell(row=linha_total, column=2).number_format = 'R$ #,##0.00'

    for coluna in pagina.columns:
        letra_coluna = coluna[0].column_letter
        maior_tamanho = 0

        for celula in coluna:
            if celula.value is not None:
                if isinstance(celula.value, (int, float)) and celula.column == 2:
                    valor_texto = f"R$ {celula.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    tamanho = len(valor_texto)
                else:
                    tamanho = len(str(celula.value))

                if tamanho > maior_tamanho:
                    maior_tamanho = tamanho

        pagina.column_dimensions[letra_coluna].width = maior_tamanho + 4

    mes_nome = meses[0] if meses else "sem_mes"

    datas_validas = [r["data_obj"] for r in registros_normalizados if r["data_obj"] is not None]

    if datas_validas:
        data_base = datas_validas[0]
    else:
        data_base = datetime.now()

    ano_principal = data_base.year
    mes_principal = data_base.month

    totais_ano = {}
    totais_mes = {}

    for r in registros_normalizados:
        categoria = r["categoria"]
        valor = r["valor"]
        data_obj = r["data_obj"]

        if categoria not in totais_ano:
            totais_ano[categoria] = 0
        if categoria not in totais_mes:
            totais_mes[categoria] = 0

        if data_obj and data_obj.year == ano_principal:
            totais_ano[categoria] += valor

            if data_obj.month == mes_principal:
                totais_mes[categoria] += valor

    criar_dashboard(workbook, totais_mes, totais_ano, mes_nome, ano_principal)

    pasta_documentos = os.path.expanduser("~/Documentos")
    os.makedirs(pasta_documentos, exist_ok=True)

    nome_arquivo = f"despesas_{mes_nome}.xlsx"
    caminho_completo = os.path.join(pasta_documentos, nome_arquivo)

    workbook.save(caminho_completo)

    return caminho_completo